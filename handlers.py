import logging
import os
import openpyxl
import re
import calendar
import datetime as dt  # Импортируем модуль datetime под псевдонимом dt
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict
from io import BytesIO
from telegram import Update
from telegram.ext import ContextTypes, CallbackContext
from config import *
from database import SQLiteDatabase, StatsCache
from keyboards import *

logger = logging.getLogger(__name__)
db = SQLiteDatabase()
stats_cache = StatsCache(ttl=1800)  # 30 минут TTL

# Компактное логирование действий пользователя
def log_action(user_id: str, action: str, data: dict = None, level: str = "INFO"):
    """Логирует действие пользователя в компактном формате"""
    message = f"USER {user_id}: {action}"
    if data:
        # Фильтруем длинные значения и чувствительные данные
        filtered_data = {}
        for k, v in data.items():
            if v is None:
                continue
            if isinstance(v, str) and len(v) > 50:
                filtered_data[k] = v[:50] + '...'
            elif isinstance(v, list):
                filtered_data[k] = f"list[{len(v)}]"
            else:
                filtered_data[k] = v
        message += f" | {filtered_data}"

    if level == "DEBUG":
        logger.debug(message)
    elif level == "WARNING":
        logger.warning(message)
    elif level == "ERROR":
        logger.error(message)
    else:
        logger.info(message)

def get_days_in_month(month, year):
    """Возвращает количество дней в указанном месяце с учетом високосных годов"""
    return calendar.monthrange(year, month)[1]

def generate_day_keyboard(month, year):
    """Генерирует клавиатуру с числами месяца"""
    days_count = get_days_in_month(month, year)
    buttons = [str(i) for i in range(1, days_count + 1)]

    # Разбиваем на ряды по 7 кнопок
    keyboard = []
    row = []
    for day in buttons:
        row.append(day)
        if len(row) == 7:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)

    keyboard.append(["Назад"])
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def validate_date(date_str):
    """Проверка корректности формата даты"""
    try:
        dt.datetime.strptime(date_str, "%d.%m.%Y")
        return True
    except ValueError:
        return False

def sanitize_filename(filename):
    """Удаление опасных символов из имени файла"""
    filename = filename.replace(' ', '_')
    return re.sub(r'[^\w_.)(-]', '', filename)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработчик команды /start"""
    try:
        user_id = str(update.message.from_user.id)

        # Очищаем временные данные
        context.user_data.clear()

        # Удаляем существующие задания для этого чата (только если job_queue доступен)
        chat_id = update.effective_chat.id

        if context.job_queue:
            current_jobs = context.job_queue.get_jobs_by_name(str(chat_id))

            # Защита от флуда заданиями
            if len(current_jobs) > 10:
                logger.warning(f"Too many jobs for chat {chat_id}, removing oldest")
                for job in current_jobs[10:]:
                    job.schedule_removal()

            # Установка напоминаний
            try:
                context.job_queue.run_daily(
                    daily_reminder,
                    time=REMINDER_TIME,
                    days=tuple(range(7)),
                    chat_id=chat_id,
                    data=user_id,
                    name=str(chat_id)
                )
                context.chat_data['job'] = True
                logger.info(f"Напоминание установлено для {user_id}")
            except Exception as e:
                logger.error(f"Ошибка установки напоминания: {e}")
        else:
            logger.warning("Job queue not available during /start command")

        await update.message.reply_text("Привет! Я твой ассистент по учету работ. Выбери категорию:",
                                      reply_markup=main_keyboard())
        return States.SELECTING_WORK
    except Exception as e:
        logger.error(f"Ошибка в функции start: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка при запуске, попробуйте позже")
        return States.SELECTING_WORK

async def handle_work_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка выбора категории работы"""
    try:
        text = update.message.text.strip().lower()
        user_data = context.user_data

        handlers = {
            "душевые": ("shower", States.SHOWER_WORK),
            "зеркала": ("mirror", States.MIRROR_WORK),
            "другая работа": ("other", States.OTHER_WORK),
            "выгрузить отчет": generate_excel,
            "удалить последнюю": delete_last,
            "просмотреть работы": view_entries,
            "статистика": show_stats,
            "⚙️ настройки": settings_menu,
            "добавить за прошлую дату": handle_past_date
        }

        # Обработка новой кнопки в категориях
        if text == "добавить за прошлую дату":
            category = user_data.get("category")
            if category:
                user_data["date_selection_source"] = category
                await update.message.reply_text("Выбери дату:", reply_markup=date_selection_keyboard())
                return States.SELECTING_DATE

        matched_handler = next((v for k, v in handlers.items() if k in text), None)
        if matched_handler:
            if isinstance(matched_handler, tuple):
                user_data["category"] = matched_handler[0]
                group_started = "current_works" in user_data
                await update.message.reply_text(
                    f"Выбери вид работы ({text}):",
                    reply_markup=work_keyboard(matched_handler[0], group_started)
                )
                return matched_handler[1]
            else:
                return await matched_handler(update, context)

        await update.message.reply_text("Пожалуйста, выбери вариант из меню", reply_markup=main_keyboard())
        return States.SELECTING_WORK
    except Exception as e:
        logger.error(f"Ошибка при выборе работы: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте позже", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def handle_past_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка кнопки 'Добавить за прошлую дату'"""
    try:
        user_data = context.user_data
        user_id = str(update.message.from_user.id)

        # Сохраняем источник запроса
        category = user_data.get("category")
        if not category:
            logger.warning(f"USER {user_id}: Категория не выбрана при добавлении за прошлую дату")
            await update.message.reply_text("Сначала выбери категорию работы", reply_markup=main_keyboard())
            return States.SELECTING_WORK

        user_data["date_selection_source"] = category
        await update.message.reply_text("Выбети дату:", reply_markup=date_selection_keyboard())
        return States.SELECTING_DATE
    except Exception as e:
        logger.error(f"Ошибка при обработке добавления за прошлую дату: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте снова", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def handle_work(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка выбора конкретной работы"""
    try:
        # Получаем информацию о пользователе
        user_id = str(update.message.from_user.id)
        user_name = update.message.from_user.full_name or f"id{user_id}"
        text = update.message.text.strip()

        # Логируем начало обработки
        logger.info(f"USER {user_id} ({user_name}): Начало обработки выбора работы. "
                    f"Ввод: '{text}', Текущие данные: {context.user_data}")

        # Основная логика обработки
        user_data = context.user_data

        if text == "Назад":
            logger.info(f"USER {user_id}: Выбрана кнопка 'Назад'. Возврат к выбору категории")
            await update.message.reply_text("Выбери категорию:", reply_markup=main_keyboard())
            return States.SELECTING_WORK

        if text == "Добавить за прошлую дату":
            logger.info(f"USER {user_id}: Запрошено добавление работы за прошлую дату")
            user_data["date_selection_source"] = user_data.get("category", "other")
            await update.message.reply_text("Выбери дату:", reply_markup=date_selection_keyboard())
            return States.SELECTING_DATE

        category = user_data.get("category")
        if not category:
            logger.warning(f"USER {user_id}: Категория не выбрана! Текущие данные: {user_data}")
            await update.message.reply_text("Ошибка: категория не выбрана", reply_markup=main_keyboard())
            return States.SELECTING_WORK

        # Обработка в зависимости от категории
        if category == "shower":
            logger.info(f"USER {user_id}: Выбрана душевая работа: '{text}'")
            user_data["shower_work"] = text
            await update.message.reply_text(
                "Добавить доп.услугу?", reply_markup=additional_services_keyboard()
            )
            return States.ADDITIONAL_SERVICES

        elif category == "mirror":
            logger.info(f"USER {user_id}: Выбрана работа с зеркалом: '{text}'")
            # ИСПРАВЛЕННАЯ ЛОГИКА ДЛЯ КНОПКИ "НАВЕС"
            if text.lower() == "навес":
                work_name = "Зеркало навес"
            elif text in ["Обычное с подсветкой", "Большое с подсветкой", "В сборной раме"]:
                work_name = f"Зеркало {text}"
            else:
                work_name = f"Зеркало {text}"  # Добавляем "Зеркало" для всех других вариантов

            user_data["mirror_work_base"] = work_name
            await update.message.reply_text("Укажи количество:", reply_markup=mirror_quantity_keyboard())
            return States.MIRROR_QUANTITY

        else:  # Другие работы
            if text == "Ввести работу вручную":
                logger.info(f"USER {user_id}: Запрошен ручной ввод работы")
                user_data["manual_input"] = True
                await update.message.reply_text(
                    "Введи название работы:",
                    reply_markup=create_keyboard(["Отмена"], add_back=False)
                )
                return States.OTHER_WORK

            # Обработка отмены при ручном вводе
            if text == "Отмена" and "manual_input" in user_data:
                logger.info(f"USER {user_id}: Отмена ручного ввода")
                del user_data["manual_input"]
                group_started = "current_works" in user_data
                await update.message.reply_text(
                    "Выбери действие:",
                    reply_markup=work_keyboard("other", group_started)
                )
                return States.OTHER_WORK

            # Обработка введенной вручную работы
            if "manual_input" in user_data:
                logger.info(f"USER {user_id}: Добавлена ручная работа: '{text}'")
                if "current_works" not in user_data:
                    user_data["current_works"] = []
                user_data["current_works"].append(text)
                del user_data["manual_input"]

                # Если это первая работа - запрашиваем адрес
                if len(user_data["current_works"]) == 1:
                    return await request_address(update, context)
                else:
                    logger.info(f"USER {user_id}: Работа добавлена в группу. "
                                f"Текущие работы: {user_data['current_works']}")
                    await update.message.reply_text(
                        "✅ Работа добавлена в группу!\n\nДобавить еще работу?",
                        reply_markup=add_more_keyboard()
                    )
                    return States.ADD_MORE_WORK

        # Если ни одно условие не сработало
        group_started = "current_works" in user_data
        logger.warning(f"USER {user_id}: Необработанный ввод в состоянии выбора работы. "
                       f"Ввод: '{text}', Категория: {category}, Данные: {user_data}")
        await update.message.reply_text("Пожалуйста, выбери вариант из меню",
                                      reply_markup=work_keyboard(category, group_started))
        return States.OTHER_WORK if category == "other" else States.SHOWER_WORK

    except Exception as e:
        # Детальное логирование исключений
        logger.error(f"USER {user_id}: Критическая ошибка в обработке работы! "
                    f"Сообщение: {update.message.text}, "
                    f"Данные: {context.user_data}, "
                    f"Ошибка: {e}",
                    exc_info=True)
        await update.message.reply_text("Произошла критическая ошибка, попробуйте снова", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def handle_additional(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка дополнительных услуг"""
    try:
        text = update.message.text.strip()
        user_data = context.user_data

        if text == "Назад":
            group_started = "current_works" in user_data
            await update.message.reply_text(
                "Выбери вид работы:", reply_markup=work_keyboard(user_data["category"], group_started)
            )
            return States.SHOWER_WORK if user_data["category"] == "shower" else States.MIRROR_WORK

        full_work = user_data["shower_work"]
        if text != "Пропустить":
            full_work += f", {text}"

        if "current_works" not in user_data:
            user_data["current_works"] = []
        user_data["current_works"].append(full_work)

        # Если это первая работа в группе - запрашиваем адрес
        if len(user_data["current_works"]) == 1:
            return await request_address(update, context)
        else:
            # Для последующих работ сразу переходим к вопросу о добавлении еще
            await update.message.reply_text(
                "✅ Работа добавлена в группу!\n\nДобавить еще работу?",
                reply_markup=add_more_keyboard()
            )
            return States.ADD_MORE_WORK
    except Exception as e:
        logger.error(f"Ошибка при обработке доп.услуг: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте снова", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def handle_mirror_quantity(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка количества зеркал"""
    try:
        text = update.message.text.strip()
        user_data = context.user_data

        if text == "Назад":
            group_started = "current_works" in user_data
            await update.message.reply_text(
                "Выбери вид работы с зеркалом:", reply_markup=work_keyboard("mirror", group_started)
            )
            return States.MIRROR_WORK

        work_name = user_data["mirror_work_base"]
        if text != "Пропустить":
            try:
                quantity = int(text)
                work_name = f"{work_name} (x{quantity})"
            except ValueError:
                await update.message.reply_text("❌ Введи число!", reply_markup=mirror_quantity_keyboard())
                return States.MIRROR_QUANTITY

        if "current_works" not in user_data:
            user_data["current_works"] = []
        user_data["current_works"].append(work_name)

        # Если это первая работа в группе - запрашиваем адрес
        if len(user_data["current_works"]) == 1:
            return await request_address(update, context)
        else:
            # Для последующих работ сразу переходим к вопросу о добавлении еще
            await update.message.reply_text(
                "✅ Работа добавена в группу!\n\nДобавить еще работу?",
                reply_markup=add_more_keyboard()
            )
            return States.ADD_MORE_WORK
    except Exception as e:
        logger.error(f"Ошибка при обработке количества зеркал: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте снова", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def request_address(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Запрос адреса"""
    try:
        await update.message.reply_text(
            "📬 Введи адрес (или 'Пропустить'):", reply_markup=create_keyboard(["Пропустить"], add_back=False)
        )
        return States.ADD_ADDRESS
    except Exception as e:
        logger.error(f"Ошибка при запросе адреса: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте снова", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def handle_address(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка адреса"""
    try:
        user_data = context.user_data
        user_data["address"] = "" if update.message.text.strip().lower() == "пропустить" else update.message.text.strip()

        await update.message.reply_text(
            "💬 Введи комментарий (или 'Пропустить'):", reply_markup=create_keyboard(["Пропустить"], add_back=False)
        )
        return States.ADD_COMMENT
    except Exception as e:
        logger.error(f"Ошибка при обработке адреса: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте снова", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def handle_comment(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка комментария"""
    try:
        user_data = context.user_data
        user_data["comment"] = "" if update.message.text.strip().lower() == "пропустить" else update.message.text.strip()

        selected_date = user_data.get("selected_date", dt.datetime.now(MOSCOW_TZ).strftime("%d.%m.%Y"))
        address = user_data.get("address", "")
        work = user_data["current_works"][-1]

        message = (
            f"✅ Работа добавлена!\nДата: {selected_date}\n"
            f"Адрес: {address or 'не указан'}\n"
            f"Работа: {work}\nКомментарий: {user_data['comment'] or 'нет'}\n\n"
            "Добавить еще работу?"
        )

        await update.message.reply_text(message, reply_markup=add_more_keyboard())
        return States.ADD_MORE_WORK
    except Exception as e:
        logger.error(f"Ошибка при обработке комментария: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте снова", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def handle_add_more(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка добавления дополнительных работ"""
    try:
        text = update.message.text.strip()
        user_data = context.user_data
        user_id = str(update.message.from_user.id)

        if text == "Завершить":
            new_entry = {
                "date": user_data.get("selected_date", dt.datetime.now(MOSCOW_TZ).strftime("%d.%m.%Y")),
                "works": user_data["current_works"].copy(),
                "comment": user_data.get("comment", ""),
                "address": user_data.get("address", "")
            }

            # Сохраняем в базе данных
            entry_id = db.add_entry(user_id, new_entry)
            if entry_id:
                # Формируем ответ с перечислением всех работ
                works_list = "\n".join([f"- {work}" for work in user_data["current_works"]])
                await update.message.reply_text(
                    f"✅ Группа работ сохранена!\nДата: {new_entry['date']}\n"
                    f"Адрес: {new_entry['address'] or 'не указан'}\n"
                    f"Комментарий: {new_entry['comment'] or 'нет'}\n"
                    f"Работы:\n{works_list}",
                    reply_markup=main_keyboard()
                )

                # Полная очистка временных данных
                keys_to_remove = [
                    "selected_date", "current_works", "address", "comment",
                    "category", "shower_work", "mirror_work_base", "manual_input",
                    "date_month_year"
                ]
                for key in keys_to_remove:
                    if key in user_data:
                        del user_data[key]

                # Инвалидация кэша статистики
                stats_cache.invalidate(user_id)
                return States.SELECTING_WORK
            else:
                await update.message.reply_text("❌ Ошибка при сохранении группы работ", reply_markup=main_keyboard())
                return States.SELECTING_WORK

        elif text == "Добавить еще работу":
            # Очищаем только данные конкретной работы
            for key in ["shower_work", "additional_service", "mirror_work_base", "manual_input"]:
                if key in user_data:
                    del user_data[key]

            await update.message.reply_text("Выбери категорию для следующей работы:", reply_markup=main_keyboard())
            return States.SELECTING_WORK

    except Exception as e:
        logger.error(f"Ошибка при добавлении дополнительных работ: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте снова", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def handle_date_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка выбора даты"""
    try:
        text = update.message.text.strip()
        user_data = context.user_data
        now = dt.datetime.now(MOSCOW_TZ)

        if text == "Отмена":
            await update.message.reply_text("Действие отменено", reply_markup=main_keyboard())
            return States.SELECTING_WORK

        selected_date = None
        if text == "Сегодня":
            selected_date = now.strftime("%d.%m.%Y")
        elif text == "Вчера":
            selected_date = (now - dt.timedelta(days=1)).strftime("%d.%m.%Y")
        elif text == "Позавчера":
            selected_date = (now - dt.timedelta(days=2)).strftime("%d.%m.%Y")
        elif text == "Текущий месяц":
            # Определяем текущий месяц и год
            current_month = now.month
            current_year = now.year
            # Генерируем клавиатуру с числами месяца
            keyboard = generate_day_keyboard(current_month, current_year)
            await update.message.reply_text(
                "Выбери число текущего месяца:",
                reply_markup=keyboard
            )
            user_data["date_month_year"] = (current_month, current_year)
            return States.SELECTING_DATE
        elif text == "Предыдущий месяц":
            # Вычисляем предыдущий месяц
            if now.month == 1:
                prev_month = 12
                prev_year = now.year - 1
            else:
                prev_month = now.month - 1
                prev_year = now.year

            # Генерируем клавиатуру с числами месяца
            keyboard = generate_day_keyboard(prev_month, prev_year)
            await update.message.reply_text(
                "Выбери число предыдущего месяца:",
                reply_markup=keyboard
            )
            user_data["date_month_year"] = (prev_month, prev_year)
            return States.SELECTING_DATE
        elif text == "Назад":
            # Возвращаемся к выбору даты
            await update.message.reply_text("Выбери дату:", reply_markup=date_selection_keyboard())
            return States.SELECTING_DATE
        else:
            # Обработка ввода числа (дня месяца)
            if text.isdigit() and "date_month_year" in user_data:
                day = int(text)
                month, year = user_data["date_month_year"]

                # Проверяем корректность дня
                days_in_month = get_days_in_month(month, year)
                if day < 1 or day > days_in_month:
                    await update.message.reply_text(
                        f"❌ В этом месяце должно быть число от 1 до {days_in_month}",
                        reply_markup=generate_day_keyboard(month, year)
                    )
                    return States.SELECTING_DATE

                # Формируем дату
                selected_date = f"{day:02d}.{month:02d}.{year}"
                user_data["selected_date"] = selected_date

                # Удаляем временные данные месяца
                del user_data["date_month_year"]
            else:
                # Валидация введенной даты
                if not validate_date(text):
                    await update.message.reply_text(
                        "❌ Неверный формат даты. Используй ДД.ММ.ГГГГ (например, 15.06.2025)",
                        reply_markup=date_selection_keyboard()
                    )
                    return States.SELECTING_DATE

                try:
                    # Парсим с учетом временной зоны
                    date_obj = dt.datetime.strptime(text, "%d.%m.%Y").replace(tzinfo=MOSCOW_TZ)
                    if date_obj.date() > now.date():
                        await update.message.reply_text(
                            "❌ Нельзя выбрать будущую дату!",
                            reply_markup=date_selection_keyboard()
                        )
                        return States.SELECTING_DATE
                    selected_date = text
                except Exception as e:
                    await update.message.reply_text(
                        f"❌ Ошибка: {str(e)}",
                        reply_markup=date_selection_keyboard()
                    )
                    return States.SELECTING_DATE

        # Если дата была выбрана из готовых вариантов
        if selected_date:
            user_data["selected_date"] = selected_date
            source = user_data.get("date_selection_source", "other")

            if source == "shower":
                group_started = "current_works" in user_data
                await update.message.reply_text(
                    f"📅 Выбрана дата: {selected_date}. Теперь выбери вид душевой:",
                    reply_markup=work_keyboard("shower", group_started)
                )
                return States.SHOWER_WORK

            if source == "mirror":
                group_started = "current_works" in user_data
                await update.message.reply_text(
                    f"📅 Выбрана дата: {selected_date}. Теперь выбери вид работы с зеркалом:",
                    reply_markup=work_keyboard("mirror", group_started)
                )
                return States.MIRROR_WORK

            group_started = "current_works" in user_data
            await update.message.reply_text(
                f"📅 Выбрана дата: {selected_date}. Теперь выбери вид работы:",
                reply_markup=main_keyboard()
            )
            return States.SELECTING_WORK

        # Если не удалось определить дату
        await update.message.reply_text("Пожалуйста, выбери вариант из меню", reply_markup=date_selection_keyboard())
        return States.SELECTING_DATE
    except Exception as e:
        logger.error(f"Ошибка при выборе даты: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте снова", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def generate_excel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Генерация Excel-отчета в памяти"""
    try:
        import time as time_lib
        user_id = str(update.message.from_user.id)
        entries = db.get_entries(user_id)

        if not entries:
            await update.message.reply_text("📭 Нет данных для отчета", reply_markup=main_keyboard())
            return States.SELECTING_WORK

        try:
            # Создаем новую книгу Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Отчет о работах"

            # Заголовки столбцов
            headers = ["Дата", "Адрес", "Вид работы", "Комментарий"]
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                cell = ws[f"{col_letter}1"]
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Сортируем записи по дате
            sorted_entries = sorted(
                entries,
                key=lambda x: dt.datetime.strptime(x['date'], "%d.%m.%Y")
            )

            # Заполняем данные
            row_num = 2
            for entry in sorted_entries:
                works_str = ", ".join(entry["works"])
                ws[f"A{row_num}"] = entry["date"]
                ws[f"B{row_num}"] = entry.get("address", "")
                ws[f"C{row_num}"] = works_str
                ws[f"D{row_num}"] = entry.get("comment", "")
                row_num += 1

            # Настраиваем ширину столбцов
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 30

            # Генерируем безопасное имя файла
            now = dt.datetime.now()
            month_name = MONTHS_GENITIVE.get(now.month, now.strftime("%B"))

            # Получаем имя пользователя
            user_name = update.message.from_user.first_name or update.message.from_user.username or f"user_{update.message.from_user.id}"
            safe_user_name = sanitize_filename(user_name)[:20]  # Ограничиваем длину

            filename = f"отчёт_{month_name}_{now.year}_{safe_user_name}.xlsx"
            safe_filename = sanitize_filename(filename)

            # Сохраняем в буфер памяти
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Отправляем файл из памяти
            await update.message.reply_document(
                document=buffer,
                filename=safe_filename,
                caption="📊 Отчет о работах",
                reply_markup=main_keyboard()
            )

            return States.SELECTING_WORK

        except Exception as e:
            logger.error(f"Ошибка генерации Excel: {e}", exc_info=True)
            await update.message.reply_text("⚠️ Ошибка при создании отчета", reply_markup=main_keyboard())
            return States.SELECTING_WORK
    except Exception as e:
        logger.error(f"Ошибка в функции генерации Excel: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка при создании отчета", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def delete_last(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Удаление последней записи"""
    try:
        user_data = context.user_data
        user_id = str(update.message.from_user.id)
        last_entry = db.get_last_entry(user_id)

        if not last_entry:
            await update.message.reply_text("❌ Нет записей для удаления", reply_markup=main_keyboard())
            return States.SELECTING_WORK

        # Сохраняем ID для последующего удаления
        user_data["pending_delete_id"] = last_entry["id"]

        # Формируем сообщение с деталями записи
        works_list = "\n".join([f"- {work}" for work in last_entry["works"]])
        message = (
            f"🗑️ Вы уверены, что хотите удалить последнюю запись?\n\n"
            f"Дата: {last_entry['date']}\n"
            f"Адрес: {last_entry.get('address', '')}\n"
            f"Комментарий: {last_entry.get('comment', '')}\n"
            f"Работы:\n{works_list}"
        )

        await update.message.reply_text(message, reply_markup=confirm_keyboard())
        return States.CONFIRM_DELETE_LAST
    except Exception as e:
        logger.error(f"Ошибка при удалении последней записи: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте позже", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def handle_confirm_delete_last(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Подтверждение удаления последней записи"""
    try:
        text = update.message.text.strip()
        user_data = context.user_data
        user_id = str(update.message.from_user.id)

        if "✅ Да, удалить" in text:
            entry_id = user_data.get("pending_delete_id")
            if entry_id:
                success = db.delete_entry(entry_id, user_id)
                if success:
                    await update.message.reply_text("✅ Последняя запись успешно удалена!", reply_markup=main_keyboard())
                    # Инвалидация кэша статистики
                    stats_cache.invalidate(user_id)
                else:
                    await update.message.reply_text("❌ Ошибка при удалении записи", reply_markup=main_keyboard())
            else:
                await update.message.reply_text("❌ Не найдена запись для удаления", reply_markup=main_keyboard())

            # Очищаем ID удаления
            if "pending_delete_id" in user_data:
                del user_data["pending_delete_id"]

            return States.SELECTING_WORK

        await update.message.reply_text("❌ Удление отменено", reply_markup=main_keyboard())
        return States.SELECTING_WORK
    except Exception as e:
        logger.error(f"Ошибка при подтверждении удаления последней записи: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте позже", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def view_entries(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Просмотр всех записей"""
    try:
        user_data = context.user_data
        user_id = str(update.message.from_user.id)
        entries = db.get_entries(user_id)

        if not entries:
            await update.message.reply_text("📭 Нет сохраненных записей", reply_markup=main_keyboard())
            return States.SELECTING_WORK

        response = "📋 Список всех работ:\n\n"
        for i, entry in enumerate(entries, 1):
            response += f"{i}. 📅 {entry['date']}\n"
            response += f"   📍 Адрес: {entry.get('address', 'не указан')}\n"
            response += f"   💬 Комментарий: {entry.get('comment', 'нет')}\n"
            response += "   🔧 Работы:\n"

            for j, work in enumerate(entry["works"], 1):
                response += f"      {j}. {work}\n"
            response += "\n"

        user_data["viewing_entries"] = entries
        await update.message.reply_text(response, reply_markup=view_entries_keyboard())
        return States.VIEWING_ENTRIES
    except Exception as e:
        logger.error(f"Ошибка при просмотре записей: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте позже", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def handle_view_entries(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка действий при просмотре записей"""
    try:
        text = update.message.text.strip()
        user_data = context.user_data

        if text == "Назад":
            await update.message.reply_text("Главное меню", reply_markup=main_keyboard())
            return States.SELECTING_WORK

        if text == "Удалить запись":
            await update.message.reply_text("Введи номер записи для удаления (или 'Отмена'):",
                                           reply_markup=create_keyboard(["Отмена"]))
            return States.DELETING_ENTRY

        await update.message.reply_text("Используй кнопки меню", reply_markup=view_entries_keyboard())
        return States.VIEWING_ENTRIES
    except Exception as e:
        logger.error(f"Ошибка при обработке просмотра записей: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте позже", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def handle_delete_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка удаления записи"""
    try:
        text = update.message.text.strip().lower()
        user_data = context.user_data
        user_id = str(update.message.from_user.id)

        if text == "отмена":
            await update.message.reply_text("Отмена удаления", reply_markup=view_entries_keyboard())
            return States.VIEWING_ENTRIES

        try:
            index = int(text) - 1
            entries = user_data.get("viewing_entries", [])
            if index < 0 or index >= len(entries):
                raise ValueError("Неверный индекс")

            # Сохраняем ID записи для подтверждения
            entry = entries[index]
            user_data["pending_delete_id"] = entry["id"]

            # Формируем сообщение с деталями записи
            works_list = "\n".join([f"- {work}" for work in entry["works"]])
            message = (
                f"🗑️ Вы уверены, что хотите удалить эту запись?\n\n"
                f"Дата: {entry['date']}\n"
                f"Адрес: {entry.get('address', '')}\n"
                f"Комментарий: {entry.get('comment', '')}\n"
                f"Работы:\n{works_list}"
            )

            await update.message.reply_text(message, reply_markup=confirm_keyboard())
            return States.CONFIRM_DELETE_ENTRY

        except ValueError:
            await update.message.reply_text("❌ Неверный номер записи (введи число)",
                                           reply_markup=create_keyboard(["Отмена"]))
            return States.DELETING_ENTRY
    except Exception as e:
        logger.error(f"Ошибка при удалении записи: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте позже", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def handle_confirm_delete_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Подтверждение удаления записи"""
    try:
        text = update.message.text.strip()
        user_data = context.user_data
        user_id = str(update.message.from_user.id)

        if "✅ Да, удалить" in text:
            entry_id = user_data.get("pending_delete_id")
            if entry_id:
                success = db.delete_entry(entry_id, user_id)
                if success:
                    await update.message.reply_text("✅ Запись успешно удалена!", reply_markup=main_keyboard())
                    # Инвалидация кэша статистики
                    stats_cache.invalidate(user_id)
                else:
                    await update.message.reply_text("❌ Ошибка при удалении записи", reply_markup=main_keyboard())
            else:
                await update.message.reply_text("❌ Не найдена запись для удаления", reply_markup=main_keyboard())

            # Очищаем временные данные
            keys = ["pending_delete_id", "viewing_entries"]
            for key in keys:
                if key in user_data:
                    del user_data[key]

            return States.SELECTING_WORK

        await update.message.reply_text("❌ Удаление отменено", reply_markup=main_keyboard())
        return States.SELECTING_WORK
    except Exception as e:
        logger.error(f"Ошибка при подтверждении удаления записи: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте позже", reply_markup=main_keyboard())
        return States.SELECTING_WORK

def calculate_stats(user_id: str) -> dict:
    """Расчет статистики для пользователя"""
    try:
        now = dt.datetime.now(MOSCOW_TZ)
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        start_date = month_start.strftime("%d.%m.%Y")
        end_date = now.strftime("%d.%m.%Y")

        entries = db.get_entries(user_id, (start_date, end_date))

        stats = {"total_groups": 0, "total_works": 0, "categories": defaultdict(int), "works": defaultdict(int)}

        for entry in entries:
            stats["total_groups"] += 1
            stats["total_works"] += len(entry['works'])

            for work in entry['works']:
                stats["works"][work] += 1
                if any(keyword in work.lower() for keyword in ["душ", "распашка", "фикс"]):
                    stats["categories"]["Душевые"] += 1
                elif "зеркал" in work.lower():
                    stats["categories"]["Зеркала"] += 1
                else:
                    stats["categories"]["Другие работы"] += 1

        return stats
    except Exception as e:
        logger.error(f"Ошибка при расчете статистики: {e}", exc_info=True)
        return {}

async def show_stats(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Отображение статистики"""
    try:
        user_id = str(update.message.from_user.id)
        stats = stats_cache.get(user_id, calculate_stats)

        if not stats or not stats.get("total_groups", 0):
            await update.message.reply_text("📭 Нет данных для статистики", reply_markup=main_keyboard())
            return States.SELECTING_WORK

        month_name = MONTHS_GENITIVE.get(dt.datetime.now().month, "")
        response = (
            f"📊 Статистика за {month_name}:\n"
            f"• Групп работ: {stats['total_groups']}\n"
            f"• Всего работ: {stats['total_works']}\n\n"
            "📋 По категориям:\n"
        )

        for category, count in stats["categories"].items():
            response += f"  - {category}: {count}\n"

        response += "\n🏆 Топ работ:\n"
        top_works = sorted(stats["works"].items(), key=lambda x: x[1], reverse=True)[:5]
        for i, (work, count) in enumerate(top_works, 1):
            response += f"  {i}. {work}: {count}\n"

        await update.message.reply_text(response, reply_markup=main_keyboard())
        return States.SELECTING_WORK
    except Exception as e:
        logger.error(f"Ошибка при показе статистики: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте позже", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def settings_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Меню настроек"""
    try:
        user_id = str(update.message.from_user.id)
        settings = db.get_settings(user_id)

        status = "✅ Включены" if settings["reminders"] else "❌ Выключены"
        vacation = "✅ Активен" if settings["vacation_mode"] else "❌ Не активен"
        work_days = ", ".join([DAYS_NAMES[i] for i in settings["work_days"]])

        response = (
            "⚙️ Настройки:\n\n"
            f"{status} - Напоминания\n"
            f"📅 Рабочие дни: {work_days}\n"
            f"{vacation} - Режим отпуска\n\n"
            "Выбери опцию для изменения:"
        )

        await update.message.reply_text(response, reply_markup=settings_keyboard())
        return States.SETTINGS
    except Exception as e:
        logger.error(f"Ошибка при открытии настроек: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте позже", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def handle_settings(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка настроек"""
    try:
        text = update.message.text.strip()
        user_data = context.user_data
        user_id = str(update.message.from_user.id)
        settings = db.get_settings(user_id)

        if text == "Назад":
            await update.message.reply_text("Главное меню", reply_markup=main_keyboard())
            return States.SELECTING_WORK

        if text == "⏰ Напоминания Вкл/Выкл":
            settings["reminders"] = not settings["reminders"]
            status = "включены" if settings["reminders"] else "выключены"
            db.save_settings(user_id, settings)
            await update.message.reply_text(f"Напоминания теперь {status}!")
            return await settings_menu(update, context)

        if text == "🏖 Режим отпуска":
            settings["vacation_mode"] = not settings["vacation_mode"]
            status = "активен" if settings["vacation_mode"] else "не активен"
            db.save_settings(user_id, settings)
            await update.message.reply_text(f"Режим отпуска теперь {status}!")
            return await settings_menu(update, context)

        if text == "📅 Рабочие дни":
            work_days = settings["work_days"]
            keyboard = []
            row = []
            for day_name in DAYS_NAMES:
                prefix = "✅ " if DAYS_MAP[day_name] in work_days else "❌ "
                row.append(f"{prefix}{day_name}")
                if len(row) == 3:
                    keyboard.append(row)
                    row = []
            if row:
                keyboard.append(row)
            keyboard.append(["Готово"])

            await update.message.reply_text(
                "Выбери рабочие дни (отмеченные дни будут активны):",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return States.SETTING_WORK_DAYS

        return await settings_menu(update, context)
    except Exception as e:
        logger.error(f"Ошибка при обработке настроек: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте позже", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def handle_work_days(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Настройка рабочих дней"""
    try:
        text = update.message.text.strip()
        user_data = context.user_data
        user_id = str(update.message.from_user.id)
        settings = db.get_settings(user_id)

        if text == "Готово":
            work_days_str = ", ".join([DAYS_NAMES[i] for i in settings["work_days"]])
            db.save_settings(user_id, settings)
            await update.message.reply_text(f"Рабочие дни обновлены: {work_days_str}")
            return await settings_menu(update, context)

        if text.startswith("✅") or text.startswith("❌"):
            day_name = text[2:]  # Удаляем префикс (✅/❌)
            if day_name in DAYS_MAP:
                day_index = DAYS_MAP[day_name]
                if day_index in settings["work_days"]:
                    settings["work_days"].remove(day_index)
                else:
                    settings["work_days"].append(day_index)
                settings["work_days"].sort()

        work_days = settings["work_days"]
        keyboard = []
        row = []
        for day_name in DAYS_NAMES:
            prefix = "✅ " if DAYS_MAP[day_name] in work_days else "❌ "
            row.append(f"{prefix}{day_name}")
            if len(row) == 3:
                keyboard.append(row)
                row = []
        if row:
            keyboard.append(row)
        keyboard.append(["Готово"])

        await update.message.reply_text(
            "Текущий выбор рабочих дней:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )

        return States.SETTING_WORK_DAYS
    except Exception as e:
        logger.error(f"Ошибка при настройке рабочих дней: {e}", exc_info=True)
        await update.message.reply_text("Произошла ошибка, попробуйте позже", reply_markup=main_keyboard())
        return States.SELECTING_WORK

async def daily_reminder(context: CallbackContext):
    """Ежедневное напоминание"""
    try:
        user_id = context.job.data
        settings = db.get_settings(user_id)

        if not settings["reminders"] or settings["vacation_mode"]:
            return

        today = dt.datetime.now(MOSCOW_TZ).weekday()
        if today not in settings["work_days"]:
            return

        today_str = dt.datetime.now(MOSCOW_TZ).strftime("%d.%m.%Y")
        entries = db.get_entries(user_id, (today_str, today_str))
        has_entries = bool(entries)

        if not has_entries:
            try:
                await context.bot.send_message(
                    chat_id=context.job.chat_id,
                    text="⏰ Напоминание! Не забудь добавить сегодняшние работы!",
                    reply_markup=main_keyboard()
                )
            except Exception as e:
                logger.error(f"Ошибка отправки напоминания: {e}")
    except Exception as e:
        logger.error(f"Ошибка напоминания: {e}", exc_info=True)

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Отмена действия"""
    keys = [
        "selected_date", "current_works", "address", "comment",
        "category", "shower_work", "mirror_work_base", "viewing_entries",
        "manual_input", "pending_delete_id", "date_month_year"
    ]
    for key in keys:
        if key in context.user_data:
            del context.user_data[key]

    await update.message.reply_text(
        "Действие отменено. Используй /start для перезапуска.",
        reply_markup=main_keyboard()
    )
    return ConversationHandler.END
